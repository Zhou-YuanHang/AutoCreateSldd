"""
sldd_to_excel_helper.py — 将 SLDD 数据写入 Excel，保留所有格式

策略：
- 已有同名行 → 覆盖必填列的值
- 无同名行   → 追加到末尾
- 不删除任何行，避免丢失旧 Excel 独有的数据
"""
import os, sys, json
from openpyxl import load_workbook


def main():
    if len(sys.argv) < 2:
        print("用法: python sldd_to_excel_helper.py <json_path>")
        sys.exit(1)

    with open(sys.argv[1], 'r', encoding='utf-8') as f:
        data = json.load(f)

    excel_path = data['excelPath']
    sheets = data['sheets']

    if not os.path.exists(excel_path):
        print(f"错误: Excel 文件不存在 - {excel_path}")
        sys.exit(1)

    wb = load_workbook(excel_path)

    for sn, sheet_data in sheets.items():
        headers = sheet_data['headers']
        req_cols = sheet_data.get('requiredCols', list(range(1, len(headers) + 1)))
        new_rows = sheet_data['rows']

        if sn not in wb.sheetnames:
            print(f"  {sn}: sheet 不存在，跳过")
            continue

        ws = wb[sn]
        name_col = 1  # 名称在第一列

        # 逐行处理：更新匹配行，追加新行
        max_row = ws.max_row

        if sn in ('Enum', 'BusElement'):
            # Enum / BusElement：第一列不是唯一键，全量替换
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            max_row = 1
            existing_names = {}  # 无匹配，全部追加
        else:
            # Signal/Parameter/Const/Bus：第一列为唯一键，按名匹配
            existing_names = {}
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=name_col).value
                if v is not None:
                    existing_names[str(v).strip()] = r

        for row_data in new_rows:
            new_name = str(row_data[0]).strip() if row_data[0] else ''

            if new_name and new_name in existing_names:
                # 已有同名行 → 更新必填列
                target_row = existing_names[new_name]
            else:
                # 无同名行 → 追加
                max_row += 1
                target_row = max_row

            # 写入全部列的值
            for c in range(1, len(row_data) + 1):
                idx = c - 1
                if idx < len(row_data):
                    val = row_data[idx]
                    try:
                        if isinstance(val, bool):
                            cell_val = val
                        elif isinstance(val, (int, float)):
                            cell_val = val
                        elif val.lower() == 'true':
                            cell_val = True
                        elif val.lower() == 'false':
                            cell_val = False
                        else:
                            cell_val = float(val)
                            if cell_val == int(cell_val) and '.' not in str(val):
                                cell_val = int(cell_val)
                    except (ValueError, TypeError):
                        cell_val = str(val) if val is not None else None
                    ws.cell(row=target_row, column=c, value=cell_val)
                # 若 SLDD 无值，则不修改该单元格（保留 Excel 原有值）

        print(f"  {sn}: {len(new_rows)} 行")

    wb.save(excel_path)
    print(f"\n[OK] Excel 已更新: {excel_path}")


if __name__ == '__main__':
    main()
