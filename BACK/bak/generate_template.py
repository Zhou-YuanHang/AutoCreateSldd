"""
generate_template.py — 生成 SLDD 数据字典 Excel 模板
双击运行即可弹出另存为窗口，默认文件名 MCU_Template.xlsx

特性：
  - 8 个 Sheet 完整复刻（History/Signal/Parameter/Const/Enum/Bus/BusElement/Config）
  - 蓝色/浅蓝表头区分必填/选填列，浅绿数据行底色
  - 下拉列表引用 Config 与 Bus 表作为数据源，方便用户集中管理
  - 冻结首行、居中格式、宋体 11pt
"""
import os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, numbers
from openpyxl.worksheet.datavalidation import DataValidation

# ═══ 另存为弹窗 ═══
import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()
root.attributes('-topmost', True)
OUT_PATH = filedialog.asksaveasfilename(
    title="保存模板文件",
    defaultextension=".xlsx",
    filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
    initialfile="MCU_Template.xlsx"
)
root.destroy()
if not OUT_PATH:
    print("已取消")
    sys.exit(0)

wb = Workbook()

# ═══ 样式 ═══
FONT_DEEP  = Font(name="宋体", size=11, bold=True,  color="FFFFFFFF")
FONT_LIGHT = Font(name="宋体", size=11, bold=True,  color="00000000")
FONT_BLUE  = Font(name="宋体", size=11, bold=True,  color="FFFFFFFF")
FONT_ENUM  = Font(name="宋体", size=11, bold=False, color="00000000")
FONT_DATA  = Font(name="宋体", size=11, bold=False, color="00000000")

FILL_DEEP  = PatternFill(patternType="solid", fgColor="FF1F4E78")
FILL_LIGHT = PatternFill(patternType="solid", fgColor="FFD9E2F3")
FILL_BLUE  = PatternFill(patternType="solid", fgColor="FF5B9BD5")
FILL_GREEN = PatternFill(patternType="solid", fgColor="FFE2F0D9")
FILL_WHITE = PatternFill(patternType="solid", fgColor="FFFFFFFF")

ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=False)
BORDER = Border()
STR_NONE = "None"

MAX_ROWS = 500  # 下拉列表覆盖行数


def set_cols(ws, widths):
    for col, w in widths:
        ws.column_dimensions[col].width = w


def freeze(ws, cell):
    ws.freeze_panes = cell


def hdr(ws, headers, fills, fonts):
    for i, (h, fl, fn) in enumerate(zip(headers, fills, fonts), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = fn, fl, ALIGN, BORDER


def drow(ws, row, values, fill=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font, c.alignment, c.border = FONT_DATA, ALIGN, BORDER
        c.number_format = numbers.FORMAT_GENERAL
        if fill:
            c.fill = fill


def add_dv(ws, col_letter, formula, prompt=None):
    """为整列添加内联下拉列表（逗号分隔字符串）"""
    dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    dv.error = "请从下拉列表中选择"
    dv.errorTitle = "无效输入"
    if prompt:
        dv.prompt = prompt
        dv.promptTitle = "提示"
        dv.showPromptMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{MAX_ROWS}")


def add_dv_ref(ws, col_letter, ref, prompt=None):
    """为整列添加引用式下拉列表（如 Config!$A$2:$A$3）"""
    dv = DataValidation(type="list", formula1=ref, allow_blank=True)
    dv.error = "请从下拉列表中选择"
    dv.errorTitle = "无效输入"
    if prompt:
        dv.prompt = prompt
        dv.promptTitle = "提示"
        dv.showPromptMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{MAX_ROWS}")


# ═══════════════════  Sheet 1: History  ═══════════════════
ws = wb.active
ws.title = "History"
freeze(ws, "A2")
set_cols(ws, [("A", 9.25), ("B", 37.875), ("C", 16.0), ("D", 61.375)])
hdr(ws,
    ["Version", "ChangeDate", "Changer", "Content"],
    [FILL_BLUE]*4, [FONT_BLUE]*4)
drow(ws, 2, ["V1.0", "2026-04-08", "AutoCreateSldd", "模板初始化"])
drow(ws, 3, [
    "说明",
    "蓝色表头为必填列，浅蓝色表头为选填列",
    "",
    "模板已内置示例数据与 Config 下拉列表，可直接改名改值后使用"
])

# ═══════════════════  Sheet 2: Signal  ═══════════════════
ws = wb.create_sheet("Signal")
freeze(ws, "A2")
set_cols(ws, [
    ("A", 19.375), ("B", 10.375), ("C", 8.125), ("D", 22.875),
    ("E", 14.875), ("F", 15.375), ("G", 21.5),
    ("I", 17.125), ("J", 5.375), ("L", 5.875), ("M", 12.875),
])
HD_SIG = ["VariableName","Package","Object","CustomStorageClass",
          "DataType","InitialValue","HeaderFile","DefinitionFile",
          "Description","Min","Max","Unit","Dimensions","Complexity"]
SIG_FILLS = [FILL_DEEP]*5 + [FILL_LIGHT] + [FILL_DEEP]*2 + [FILL_LIGHT]*6
SIG_FONTS = [FONT_DEEP]*5 + [FONT_LIGHT] + [FONT_DEEP]*2 + [FONT_LIGHT]*6
hdr(ws, HD_SIG, SIG_FILLS, SIG_FONTS)

# 下拉列表（引用 Config 数据源）
add_dv_ref(ws, "B", "Config!$A$2:$A$3",     "选择包名")
add_dv_ref(ws, "C", "Config!$B$2:$B$2",     "选择 Signal")
add_dv_ref(ws, "D", "Config!$C$2:$C$7",     "选择存储类")
add_dv_ref(ws, "E", "Config!$D$2:$D$9",     "选择数据类型")
add_dv_ref(ws, "L", "Config!$E$2:$E$9",     "选择单位")
add_dv_ref(ws, "N", "Config!$G$2:$G$3",     "real 或 complex")

drow(ws, 2, ["M_Fault_StatusBus","myPackage","Signal","YS_SelectRAMSignal",
     "Bus: FaultBus","None","MCU_Fault_Signals.h","MCU_Fault_Signals.c",
     "故障状态信号总线","None","None","-","1","real"], FILL_GREEN)
drow(ws, 3, ["M_Ctrl_Position_m","myPackage","Signal","ExportedGlobal",
     "single","None","MCU_Ctrl_Signals.h","MCU_Ctrl_Signals.c",
     "位置反馈信号","0","100","m","1","real"], FILL_GREEN)

# ═══════════════════  Sheet 3: Parameter  ═══════════════════
ws = wb.create_sheet("Parameter")
freeze(ws, "A2")
set_cols(ws, [
    ("A", 21.5), ("B", 10.375), ("D", 22.875),
    ("E", 10.375), ("F", 15.375), ("G", 19.375),
    ("I", 20.25), ("J", 4.75), ("K", 5.375), ("L", 5.875), ("M", 12.875),
])
PAR_FILLS = [FILL_DEEP]*8 + [FILL_LIGHT]*6
PAR_FONTS = [FONT_DEEP]*8 + [FONT_LIGHT]*6
hdr(ws, HD_SIG, PAR_FILLS, PAR_FONTS)

# 下拉列表（引用 Config 数据源）
add_dv_ref(ws, "B", "Config!$A$2:$A$3",     "选择包名")
add_dv_ref(ws, "C", "Config!$B$3:$B$3",     "选择 Parameter")
add_dv_ref(ws, "D", "Config!$C$2:$C$7",     "选择存储类")
add_dv_ref(ws, "E", "Config!$D$2:$D$9",     "选择数据类型")
add_dv_ref(ws, "L", "Config!$E$2:$E$9",     "选择单位")
add_dv_ref(ws, "N", "Config!$G$2:$G$3",     "real 或 complex")

drow(ws, 2, ["C_Fault_IaLimit_s32","myPackage","Parameter","YS_SelectRAMPara",
     "single","600","MCU_Fault_Paras.h","MCU_Fault_Paras.c",
     "A相电流故障限值设置","0","1000","A","1","real"], FILL_GREEN)
drow(ws, 3, ["Kp_SpdLoop","myPackage","Parameter","ExportedGlobal",
     "double","0.35","MCU_Ctrl_Paras.h","MCU_Ctrl_Paras.c",
     "速度环比例系数","0","10","-","1","real"], FILL_GREEN)

# ═══════════════════  Sheet 4: Const  ═══════════════════
ws = wb.create_sheet("Const")
freeze(ws, "A2")
set_cols(ws, [
    ("A", 9.375), ("B", 6.375), ("C", 10.375), ("D", 12.875),
])
hdr(ws,
    ["Name", "Value", "DataType", "HeaderFile"],
    [FILL_DEEP]*4, [FONT_DEEP]*4)

add_dv_ref(ws, "C", "Config!$D$2:$D$9", "选择数据类型")

drow(ws, 2, ["Mode", 1, "uint8", "Define.h"], FILL_GREEN)

# ═══════════════════  Sheet 5: Enum  ═══════════════════
ws = wb.create_sheet("Enum")
set_cols(ws, [("A", 9.375), ("B", 12.625), ("C", 6.375), ("D", 9.375)])
hdr(ws,
    ["EnumName", "EnumNumbers", "Value", "DataType"],
    [FILL_DEEP]*4, [FONT_ENUM]*4)

add_dv(ws, "D", "single,double,uint8,uint16,uint32,int8,int16,int32,boolean,Bus: FaultBus,Bus: CtrlBus", "选择数据类型")

drow(ws, 2, ["GEAR", "H", 0, "uint8"])
drow(ws, 3, ["GEAR", "M", 1, "uint8"])
drow(ws, 4, ["GEAR", "L", 2, "uint8"])

# ═══════════════════  Sheet 6: Bus  ═══════════════════
ws = wb.create_sheet("Bus")
freeze(ws, "A2")
set_cols(ws, [
    ("A", 9.375), ("B", 14.125), ("C", 12.875),
    ("D", 11.625), ("E", 31.625), ("F", 11.625),
])
hdr(ws,
    ["BusName","Description","HeaderFile","Alignment",
     "PreserveElementDimensions","DataScope"],
    [FILL_DEEP]*6, [FONT_DEEP]*6)

add_dv_ref(ws, "F", "Config!$H$2:$H$4", "Auto / Exported / Imported")

drow(ws, 2, ["FaultBus","故障状态总线","fault_bus.h","-1","False","Auto"], FILL_GREEN)
drow(ws, 3, ["CtrlBus","控制总线","ctrl_bus.h","-1","False","Exported"], FILL_GREEN)

# ═══════════════════  Sheet 7: BusElement  ═══════════════════
ws = wb.create_sheet("BusElement")
freeze(ws, "A2")
set_cols(ws, [
    ("A", 9.375), ("B", 14.875), ("C", 10.375),
    ("D", 12.875), ("E", 19.125), ("F", 5.875),
])
BE_FILLS = [FILL_DEEP]*4 + [FILL_LIGHT]*2
BE_FONTS = [FONT_DEEP]*4 + [FONT_LIGHT]*2
hdr(ws,
    ["BusName","ElementName","DataType","Dimensions","Description","Unit"],
    BE_FILLS, BE_FONTS)

# 下拉列表
add_dv_ref(ws, "A", "Bus!$A$2:$A$1000", "选择总线名")
add_dv_ref(ws, "C", "Config!$D$2:$D$9", "选择数据类型")
add_dv_ref(ws, "F", "Config!$E$2:$E$9", "选择单位")

drow(ws, 2, ["FaultBus","IcSensorFault","uint8","1","电流传感器故障标志","-"], FILL_GREEN)
drow(ws, 3, ["FaultBus","VoltageFault","boolean","1","母线电压故障标志","-"], FILL_GREEN)
drow(ws, 4, ["CtrlBus","Position","double","[3,1]","位置矢量","m"], FILL_GREEN)
drow(ws, 5, ["CtrlBus","Velocity","double","[3,1]","速度矢量","m/s"], FILL_GREEN)

# ═══════════════════  Sheet 8: Config  ═══════════════════
ws = wb.create_sheet("Config")
freeze(ws, "A2")
set_cols(ws, [
    ("A", 10.375), ("C", 23.75), ("D", 14.875),
    ("E", 5.875), ("F", 17.875), ("G", 12.875), ("H", 11.625),
])
hdr(ws,
    ["Package","Object","Storage class","Data Type",
     "Unit","DimensionsMode","Complexity","DataScope"],
    [FILL_BLUE]*8, [FONT_BLUE]*8)

# Config 表不设下拉列表——它本身就是其他表的数据源

drow(ws, 2, ["myPackage","Signal","YS_SelectRAMSignal","single","-","Fixed","real","Auto"])
drow(ws, 3, ["Simulink","Parameter","YS_SelectRAMPara","double","V","Variable","complex","Exported"])
drow(ws, 4, [STR_NONE,STR_NONE,"ExportedGlobal","uint8","A",STR_NONE,STR_NONE,"Imported"])
drow(ws, 5, [STR_NONE,STR_NONE,"ImportedExtern","int16","N*m",STR_NONE,STR_NONE,STR_NONE])
drow(ws, 6, [STR_NONE,STR_NONE,"ImportedExternPointer","uint16","m",STR_NONE,STR_NONE,STR_NONE])
drow(ws, 7, [STR_NONE,STR_NONE,"Auto","uint32","m/s",STR_NONE,STR_NONE,STR_NONE])
drow(ws, 8, [STR_NONE,STR_NONE,STR_NONE,"boolean","℃",STR_NONE,STR_NONE,STR_NONE])
drow(ws, 9, [STR_NONE,STR_NONE,STR_NONE,"Bus: FaultBus","deg",STR_NONE,STR_NONE,STR_NONE])

# ═══ 保存 ═══
wb.save(OUT_PATH)
print(f"[OK] 模板已生成: {OUT_PATH}")
